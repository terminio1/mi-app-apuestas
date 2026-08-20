import streamlit as st
import pandas as pd
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuración de página
st.set_page_config(page_title="Strike Analytics Pro", page_icon="🐰", layout="wide")

st.markdown("""
<style>
.stApp {
    background-color: #E6F3EA;
}
.stMetric {
    background-color: white;
    padding: 10px;
    border-radius: 10px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.05);
}
</style>
""", unsafe_allow_html=True)

# Inicializar Historial
if "historial_apuestas" not in st.session_state:
    st.session_state["historial_apuestas"] = []

# ENCABEZADO
col_head1, col_head2 = st.columns([1, 4])

with col_head1:
    st.markdown(
        """
        <div style="text-align: center; background-color: white; border-radius: 50%; padding: 15px; width: 100px; height: 100px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); margin: auto;">
            <span style="font-size: 60px;">🐰</span>
        </div>
        """,
        unsafe_allow_html=True
    )

with col_head2:
    st.markdown(
        """
        <div style="background-color: white; border-radius: 15px; padding: 15px; border: 1px solid #FFEBEB;">
            <h1 style="color: #A32A2A; margin: 0; font-size: 32px;">Strike Analytics Pro 👋</h1>
            <p style="color: #A32A2A; margin: 5px 0 0 0;">Generador Matricial, Efecto Dominó y Gestor de Historial</p>
        </div>
        """,
        unsafe_allow_html=True
    )

tab1, tab2 = st.tabs(["🎯 Generador & Simulador Dominó", "📋 Historial & Modificador Excel"])

with tab1:
    st.header("1. Ingreso de Datos (Partidos, Equipos y Cuotas)")

    col1, col2, col3 = st.columns(3)
    partidos = []

    for idx, col in enumerate([col1, col2, col3], start=1):
        with col:
            st.markdown(
                f"""
                <div style="background-color: #F3F9F6; border-radius: 10px; padding: 10px; border: 1px solid #C4E2D5;">
                    <h3 style="margin: 0;">⚽ Partido {idx}</h3>
                </div>
                """,
                unsafe_allow_html=True
            )
            
            p_local = st.text_input(f"Local P{idx}", value=f"Local {idx}", key=f"p_loc_{idx}")
            p_visitante = st.text_input(f"Visitante P{idx}", value=f"Visitante {idx}", key=f"p_vis_{idx}")
            
            c1 = st.number_input(f"Cuota 1 ({p_local})", min_value=1.01, value=2.10, step=0.05, key=f"c1_{idx}")
            cX = st.number_input(f"Cuota X (Empate)", min_value=1.01, value=3.20, step=0.05, key=f"cX_{idx}")
            c2 = st.number_input(f"Cuota 2 ({p_visitante})", min_value=1.01, value=3.50, step=0.05, key=f"c2_{idx}")
            
            prob1, probX, prob2 = 1/c1, 1/cX, 1/c2
            total_prob = prob1 + probX + prob2
            p1_real, pX_real, p2_real = (prob1/total_prob)*100, (probX/total_prob)*100, (prob2/total_prob)*100
            
            if c1 <= c2 and c1 <= cX:
                base_sugerida = f"1 ({p_local})"
                cuota_base = c1
            elif c2 <= c1 and c2 <= cX:
                base_sugerida = f"2 ({p_visitante})"
                cuota_base = c2
            else:
                base_sugerida = "X (Empate)"
                cuota_base = cX

            st.markdown(
                f"""
                <div style="background-color: #E8F5E9; border-radius: 8px; padding: 8px; margin-top: 10px; border: 1px solid #A5D6A7;">
                    <p style="margin:0; color:#2E7D32; font-weight:bold; font-size:13px;">🤖 Recomendación Base:</p>
                    <p style="margin:0; font-size:13px;"><b>{base_sugerida}</b> @ {cuota_base:.2f}</p>
                </div>
                """,
                unsafe_allow_html=True
            )

            partidos.append({
                "local": p_local, "visitante": p_visitante, "vs": f"{p_local} vs {p_visitante}",
                "c1": c1, "cX": cX, "c2": c2,
                "opcion_base": base_sugerida, "cuota_base": cuota_base,
                "p1_prob": p1_real, "pX_prob": pX_real, "p2_prob": p2_real
            })

    # BOLETOS 8 Y 9
    st.header("2. Boletos Libres (8 y 9)")
    auto_pilot = st.checkbox("🤖 Auto-Piloto IA para Boletos 8 y 9", value=True)
    
    opciones_p1 = [f"1 ({partidos[0]['local']})", "X (Empate)", f"2 ({partidos[0]['visitante']})"]
    opciones_p2 = [f"1 ({partidos[1]['local']})", "X (Empate)", f"2 ({partidos[1]['visitante']})"]
    opciones_p3 = [f"1 ({partidos[2]['local']})", "X (Empate)", f"2 ({partidos[2]['visitante']})"]

    if auto_pilot:
        b8_p1, b8_p2, b8_p3 = "X (Empate)", "X (Empate)", partidos[2]["opcion_base"]
        def op_sorpresa(p): return f"2 ({p['visitante']})" if "1" in p["opcion_base"] else f"1 ({p['local']})"
        b9_p1, b9_p2, b9_p3 = op_sorpresa(partidos[0]), op_sorpresa(partidos[1]), op_sorpresa(partidos[2])
    else:
        col_b8, col_b9 = st.columns(2)
        with col_b8:
            b8_p1 = st.selectbox("P1 (Boleto 8)", opciones_p1, index=1, key="b8_p1")
            b8_p2 = st.selectbox("P2 (Boleto 8)", opciones_p2, index=1, key="b8_p2")
            b8_p3 = st.selectbox("P3 (Boleto 8)", opciones_p3, index=0, key="b8_p3")
        with col_b9:
            b9_p1 = st.selectbox("P1 (Boleto 9)", opciones_p1, index=2, key="b9_p1")
            b9_p2 = st.selectbox("P2 (Boleto 9)", opciones_p2, index=2, key="b9_p2")
            b9_p3 = st.selectbox("P3 (Boleto 9)", opciones_p3, index=2, key="b9_p3")

    # STAKE
    st.header("3. Configuración del Stake")
    monto_por_boleto = st.number_input("Monto por Boleto (€)", min_value=1.0, value=10.0, step=1.0)
    stake_unidad = monto_por_boleto / 4.0
    inversion_total = monto_por_boleto * 9

    st.success(f"💡 Inversión Total (9 Boletos): *{inversion_total:.2f}€* ({stake_unidad:.2f}€ por apuesta individual dentro de Trixie)")

    # GENERAR MATRIZ
    if st.button("🚀 Calcular Matriz Completa y Simulador Dominó", use_container_width=True):
        boletos_data = []

        def obtener_cuota(p_dict, sel_texto):
            if "1" in sel_texto: return p_dict["c1"]
            elif "2" in sel_texto: return p_dict["c2"]
            else: return p_dict["cX"]

        def agregar_boleto(num, tipo, sel1, sel2, sel3):
            c1 = obtener_cuota(partidos[0], sel1)
            c2 = obtener_cuota(partidos[1], sel2)
            c3 = obtener_cuota(partidos[2], sel3)
            
            d12 = c1 * c2
            d13 = c1 * c3
            d23 = c2 * c3
            triple = c1 * c2 * c3
            
            pago_dobles = (d12 + d13 + d23) * stake_unidad
            pago_trixie = (d12 + d13 + d23 + triple) * stake_unidad
            
            boletos_data.append({
                "ID": num,
                "Boleto": f"Boleto {num}",
                "Estrategia": tipo,
                f"P1: {partidos[0]['vs']}": f"{sel1} ({c1:.2f})",
                f"P2: {partidos[1]['vs']}": f"{sel2} ({c2:.2f})",
                f"P3: {partidos[2]['vs']}": f"{sel3} ({c3:.2f})",
                "Cobro Dobles (€)": round(pago_dobles, 2),
                "Cobro Trixie (€)": round(pago_trixie, 2),
                "c1": c1, "c2": c2, "c3": c3,
                "sel1": sel1, "sel2": sel2, "sel3": sel3,
                "d12_val": round(d12 * stake_unidad, 2),
                "d13_val": round(d13 * stake_unidad, 2),
                "d23_val": round(d23 * stake_unidad, 2),
                "triple_val": round(triple * stake_unidad, 2)
            })

        op_b1, op_b2, op_b3 = partidos[0]["opcion_base"], partidos[1]["opcion_base"], partidos[2]["opcion_base"]
        def inv_op(p): return f"2 ({p['visitante']})" if "1" in p["opcion_base"] else f"1 ({p['local']})"

        agregar_boleto(1, "🔥 BASE PRINCIPAL", op_b1, op_b2, op_b3)
        agregar_boleto(2, "🛡️ Cobertura Empate P1", "X (Empate)", op_b2, op_b3)
        agregar_boleto(3, "🛡️ Cobertura Empate P2", op_b1, "X (Empate)", op_b3)
        agregar_boleto(4, "🛡️ Cobertura Empate P3", op_b1, op_b2, "X (Empate)")
        agregar_boleto(5, "⚡ Cobertura Sorpresa P1", inv_op(partidos[0]), op_b2, op_b3)
        agregar_boleto(6, "⚡ Cobertura Sorpresa P2", op_b1, inv_op(partidos[1]), op_b3)
        agregar_boleto(7, "⚡ Cobertura Sorpresa P3", op_b1, op_b2, inv_op(partidos[2]))
        agregar_boleto(8, "🎯 Cierre Libre 1", b8_p1, b8_p2, b8_p3)
        agregar_boleto(9, "🚀 Cierre Libre 2", b9_p1, b9_p2, b9_p3)

        st.session_state["matriz_actual"] = boletos_data
        st.session_state["inversion_actual"] = inversion_total
        st.session_state["partidos_actuales"] = partidos
        st.session_state["stake_unidad"] = stake_unidad

    if "matriz_actual" in st.session_state:
        df_matriz = pd.DataFrame(st.session_state["matriz_actual"])
        
        st.header("📋 Matriz de 9 Boletos Generada")
        cols_ocultar = ["c1", "c2", "c3", "sel1", "sel2", "sel3", "d12_val", "d13_val", "d23_val", "triple_val", "ID"]
        df_vista = df_matriz.drop(columns=[c for c in cols_ocultar if c in df_matriz.columns])
        st.dataframe(df_vista, use_container_width=True)

        # SIMULADOR Y EFECTO DOMINÓ
        st.markdown("---")
        st.header("🔮 Simulador y Desglose del Efecto Dominó")
        
        boleto_ganador_id = st.selectbox(
            "Selecciona qué Boleto acertó los 3 partidos:",
            options=[b["ID"] for b in st.session_state["matriz_actual"]],
            format_func=lambda x: f"Boleto {x} ({st.session_state['matriz_actual'][x-1]['Estrategia']})"
        )

        b_ganador = st.session_state["matriz_actual"][boleto_ganador_id - 1]
        st_unit = st.session_state.get("stake_unidad", stake_unidad)

        d12_v = b_ganador.get("d12_val", round((b_ganador["c1"] * b_ganador["c2"]) * st_unit, 2))
        d13_v = b_ganador.get("d13_val", round((b_ganador["c1"] * b_ganador["c3"]) * st_unit, 2))
        d23_v = b_ganador.get("d23_val", round((b_ganador["c2"] * b_ganador["c3"]) * st_unit, 2))
        trip_v = b_ganador.get("triple_val", round((b_ganador["c1"] * b_ganador["c2"] * b_ganador["c3"]) * st_unit, 2))
        
        st.subheader(f"📊 Desglose de Ganancias para Boleto {boleto_ganador_id}")
        
        c1_col, c2_col, c3_col, c4_col = st.columns(4)
        c1_col.metric("Doble 1 (P1 x P2)", f"{d12_v:.2f}€")
        c2_col.metric("Doble 2 (P1 x P3)", f"{d13_v:.2f}€")
        c3_col.metric("Doble 3 (P2 x P3)", f"{d23_v:.2f}€")
        c4_col.metric("Triple Trixie (P1 x P2 x P3)", f"{trip_v:.2f}€")

        suma_dobles_otros = 0.0
        detalles_rescate = []
        for b in st.session_state["matriz_actual"]:
            if b["ID"] != boleto_ganador_id:
                m1 = (b["sel1"] == b_ganador["sel1"])
                m2 = (b["sel2"] == b_ganador["sel2"])
                m3 = (b["sel3"] == b_ganador["sel3"])
                
                sum_b = 0
                if m1 and m2: sum_b += (b["c1"] * b["c2"]) * st_unit
                if m1 and m3: sum_b += (b["c1"] * b["c3"]) * st_unit
                if m2 and m3: sum_b += (b["c2"] * b["c3"]) * st_unit
                
                if sum_b > 0:
                    suma_dobles_otros += sum_b
                    detalles_rescate.append(f"• *Boleto {b['ID']}*: aporta {sum_b:.2f}€ por compartir 2 aciertos.")

        cobro_trixie_ganador = b_ganador["Cobro Trixie (€)"]
        cobro_total_simulado = cobro_trixie_ganador + suma_dobles_otros
        neto_simulado = cobro_total_simulado - st.session_state["inversion_actual"]

        st.markdown("*Efecto Dominó en Coberturas:*")
        if detalles_rescate:
            for d in detalles_rescate:
                st.markdown(d)
        else:
            st.write("Ningún otro boleto comparte 2 aciertos exactos con esta combinación.")

        res_c1, res_c2, res_c3 = st.columns(3)
        res_c1.metric("Cobro Trixie Ganador", f"{cobro_trixie_ganador:.2f}€")
        res_c2.metric("Rescate Dobles (Otros Boletos)", f"{suma_dobles_otros:.2f}€")
        res_c3.metric("🔥 GRAN TOTAL A COBRAR", f"{cobro_total_simulado:.2f}€", delta=f"{neto_simulado:+.2f}€ Neto")

        # GUARDAR EN HISTORIAL
        st.markdown("---")
        st.subheader("💾 Guardar esta Jornada en tu Historial")
        col_g1, col_g2, col_g3 = st.columns(3)
        
        with col_g1:
            nombre_jornada = st.text_input("Nombre / Fecha de la Jornada", value=f"Jornada {len(st.session_state['historial_apuestas']) + 1}")
        with col_g2:
            resultado_final = st.selectbox("Resultado de la Apuesta", ["Ganada (Victoria)", "Perdida", "Recuperación Parcial"])
        with col_g3:
            monto_cobrado_real = st.number_input("Monto Real Cobrado (€)", value=float(round(cobro_total_simulado, 2)))

        if st.button("📌 Guardar en Historial"):
            p_act = st.session_state["partidos_actuales"]
            resumen_partidos = f"{p_act[0]['vs']} | {p_act[1]['vs']} | {p_act[2]['vs']}"
            
            st.session_state["historial_apuestas"].append({
                "Fecha": datetime.now().strftime("%Y-%m-%d"),
                "Jornada": nombre_jornada,
                "Equipos / Partidos": resumen_partidos,
                "Inversión (€)": st.session_state["inversion_actual"],
                "Cobrado (€)": monto_cobrado_real,
                "Beneficio (€)": round(monto_cobrado_real - st.session_state["inversion_actual"], 2),
                "Estado": resultado_final
            })
            st.success("¡Jornada guardada correctamente!")

# PESTAÑA 2: HISTORIAL Y EXPORTACIÓN
with tab2:
    st.header("📋 Historial Unificado, Modificador y Exportación")

    if len(st.session_state["historial_apuestas"]) == 0:
        st.info("No hay jornadas registradas. Guarda una desde la Pestaña 1.")
    else:
        st.subheader("🛠️ Editar o Eliminar Jornadas Registradas")
        
        for idx_h, item in enumerate(st.session_state["historial_apuestas"]):
            c_h1, c_h2, c_h3, c_h4, c_h5, c_h6 = st.columns([1.5, 2, 3, 2, 2, 1])
            
            with c_h1:
                item["Fecha"] = st.text_input("Fecha", value=item.get("Fecha", datetime.now().strftime("%Y-%m-%d")), key=f"edit_f_{idx_h}")
            with c_h2:
                item["Jornada"] = st.text_input("Jornada", value=item["Jornada"], key=f"edit_j_{idx_h}")
            with c_h3:
                item["Equipos / Partidos"] = st.text_input("Equipos", value=item.get("Equipos / Partidos", "P1 | P2 | P3"), key=f"edit_eq_{idx_h}")
            with c_h4:
                item["Cobrado (€)"] = st.number_input("Cobrado (€)", value=float(item["Cobrado (€)"]), key=f"edit_c_{idx_h}")
                item["Beneficio (€)"] = round(item["Cobrado (€)"] - item["Inversión (€)"], 2)
            with c_h5:
                item["Estado"] = st.selectbox("Estado", ["Ganada (Victoria)", "Perdida", "Recuperación Parcial"], 
                                             index=["Ganada (Victoria)", "Perdida", "Recuperación Parcial"].index(item["Estado"]) if item["Estado"] in ["Ganada (Victoria)", "Perdida", "Recuperación Parcial"] else 0, 
                                             key=f"edit_e_{idx_h}")
            with c_h6:
                st.write("")
                st.write("")
                if st.button("🗑️", key=f"del_{idx_h}"):
                    st.session_state["historial_apuestas"].pop(idx_h)
                    st.rerun()

        st.markdown("---")
        st.subheader("📊 Vista Previa del Historial Global")
        df_historial = pd.DataFrame(st.session_state["historial_apuestas"])
        
        def colorear_estado(val):
            if "Victoria" in str(val) or "Ganada" in str(val):
                return 'background-color: #C8E6C9; color: #2E7D32; font-weight: bold;'
            elif "Perdida" in str(val):
                return 'background-color: #FFCDD2; color: #C62828; font-weight: bold;'
            else:
                return 'background-color: #FFF9C4; color: #F57F17; font-weight: bold;'

        st.dataframe(df_historial.style.map(colorear_estado, subset=["Estado"]), use_container_width=True)

        st.markdown("---")
        st.subheader("📥 Exportar a Excel Profesional")

        buffer = io.BytesIO()
        wb = pd.ExcelWriter(buffer, engine='openpyxl')

        df_historial.to_excel(wb, sheet_name='Historial Global', index=False)

        if "matriz_actual" in st.session_state:
            p_act = st.session_state["partidos_actuales"]
            st_u = st.session_state.get("stake_unidad", 2.5)
            detalles_excel = []
            for b in st.session_state["matriz_actual"]:
                d12_e = b.get("d12_val", round((b["c1"] * b["c2"]) * st_u, 2))
                d13_e = b.get("d13_val", round((b["c1"] * b["c3"]) * st_u, 2))
                d23_e = b.get("d23_val", round((b["c2"] * b["c3"]) * st_u, 2))
                trip_e = b.get("triple_val", round((b["c1"] * b["c2"] * b["c3"]) * st_u, 2))
                
                detalles_excel.append({
                    "Boleto": b["Boleto"],
                    "Estrategia": b["Estrategia"],
                    f"Partido 1 ({p_act[0]['vs']})": b["sel1"],
                    "Cuota P1": b["c1"],
                    f"Partido 2 ({p_act[1]['vs']})": b["sel2"],
                    "Cuota P2": b["c2"],
                    f"Partido 3 ({p_act[2]['vs']})": b["sel3"],
                    "Cuota P3": b["c3"],
                    "Cobro Doble 1 (P1xP2) (€)": d12_e,
                    "Cobro Doble 2 (P1xP3) (€)": d13_e,
                    "Cobro Doble 3 (P2xP3) (€)": d23_e,
                    "Cobro Triple Trixie (€)": trip_e,
                    "TOTAL TRIXIE COMPLETO (€)": round(b["Cobro Trixie (€)"], 2)
                })
            df_mat_excel = pd.DataFrame(detalles_excel)
            df_mat_excel.to_excel(wb, sheet_name='Matriz_Detallada_Boletos', index=False)

        wb.close()

        buffer.seek(0)
        workbook = openpyxl.load_workbook(buffer)

        # ESTILOS MEJORADOS (ESTÉTICA MODERNA VERDE + FILAS CEBRA)
        header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        zebra_fill = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        thin_border = Border(left=Side(style='thin', color='C8E6C9'),
                             right=Side(style='thin', color='C8E6C9'),
                             top=Side(style='thin', color='C8E6C9'),
                             bottom=Side(style='thin', color='C8E6C9'))

        for sheet in workbook.worksheets:
            # Encabezado
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Celdas y Filas Cebra
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                fill_color = zebra_fill if row_idx % 2 == 0 else white_fill
                for cell in row:
                    cell.fill = fill_color
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00 €'

            # Ancho de Columnas
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 5, 14)

        output_excel = io.BytesIO()
        workbook.save(output_excel)

        st.download_button(
            label="📊 Descargar Excel Profesional con Formato y Colores",
            data=output_excel.getvalue(),
            file_name="Strike_Analytics_Matriz_y_Historial.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
